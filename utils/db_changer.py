import requests
from bs4 import BeautifulSoup
from application.models import Article, Science

HEADERS = {'User-Agent': 'Mozilla/5.0'}
DOMAIN = "http://synergy-journal.ru/"


def link_to_gdrive():
    for a in Article.objects.filter():

        print(a.make_url_to_remote())

        r = requests.get(a.make_url_to_remote(), headers=HEADERS)  # Выполняем запрос
        bs = BeautifulSoup(r.text, 'html.parser')  # Распарс ответа
        try:
            print(bs.find('iframe')["src"])
            a.url = bs.find('iframe')["src"]
            a.save()
        except:
            pass


cyril = "абвгдеёжзийклмнопрстуфхцчшщъыьэюя".upper()


def _it_cyrilic(word):
    cyril_character = [s in cyril for s in word]

    return cyril_character.count(True) > len(word) // 2


def change_titles():
    for a in Article.objects.filter(title_eng=""):

        s = a.title

        words = s.split(" ")

        words = [x for x in words if x]

        title_rus = []
        title_eng = []

        change_language = None
        change_counter = 0

        for word in words:

            if _it_cyrilic(word):
                title_rus.append(word)
            else:
                title_eng.append(word)

            if _it_cyrilic(word) != change_language:
                change_counter = change_counter + 1
                change_language = _it_cyrilic(word)

        if change_counter == 2:
            print(" ".join(title_rus))
            print(" ".join(title_eng))

            a.title = " ".join(title_rus)
            a.title_eng = " ".join(title_eng)

            a.save()

            # if change_counter > 2:
            #     print(" ".join(title_rus))
            #     print(" ".join(title_eng))


def make_title_and_pages():
    for a in Article.objects.filter(input_data__icontains="Выходные данные: "):

        field = a.input_data

        a.magazine_title = a.input_data[field.rindex("Выходные данные: ")+17:field.index("// Синергия наук")]
        a.pages = field[field.index("− С.")+5:field.index(" − URL: ")]

        a.save()


def make_sci():
    from openpyxl import load_workbook

    wb = load_workbook(filename='data.xlsx')
    ws = wb["для Артема"]
    for row in ws:

        # print([x.value for x in row])
        row = [x.value for x in row]

        if row[-4] == None:
            continue

        url, sci = [row[-4], row[-1]]
        print(url, sci)

        number = int(str(url)[-4:])
        print(number)

        article = Article.objects.get(number=number)
        article.science = Science.objects.get(name=sci)
        #article.save()
        print(article)


def make_authours():
    from openpyxl import load_workbook

    wb = load_workbook(filename='data.xlsx')
    ws = wb["для Артема"]
    for row in ws:

        # print([x.value for x in row])
        row = [x.value for x in row]

        if row[-4] == None:
            continue

        author, title, url = [row[0], row[1], row[-2]]
        print(author, title)

        number = int(str(url)[-4:])
        print(number)

        article = Article.objects.get(number=number)
        #article.magazine_authors = author
        article.magazine_title = title
        article.save()
        print(article)





